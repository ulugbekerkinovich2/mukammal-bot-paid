import asyncio
import json
import logging
import os
import shutil
from datetime import datetime
from typing import Any, Dict, Optional

from openpyxl import Workbook, load_workbook

logger = logging.getLogger(__name__)

EXCEL_PATH = os.getenv("MANDAT_RESULTS_XLSX", "mandat_results.xlsx")

_EXTRA_LABELS = ["Imtiyoz ball", "Ijodiy ball", "CEFR ball", "Milliy sertifikat ball / Olimpiada"]

_HEADERS = [
    "entrant_id", "full_name", "lang",
    "majburiy_fan", "majburiy_togri", "majburiy_foiz", "majburiy_ball",
    "fan1_nomi", "fan1_togri", "fan1_foiz", "fan1_ball",
    "fan2_nomi", "fan2_togri", "fan2_foiz", "fan2_ball",
    "jami_togri", "jami_savol", "jami_foiz",
    "imtiyoz_ball", "ijodiy_ball", "cefr_ball", "milliy_sertifikat_ball",
    "umumiy_ball",
    "updated_at",
]

_OLD_HEADERS = [
    "entrant_id", "full_name", "lang",
    "subjects_json", "scores_json", "extra_scores_json", "total_ball",
    "updated_at",
]

_LOCK = asyncio.Lock()


def _fmt_pct(value: float, decimals: int = 1) -> str:
    return f"{value:.{decimals}f}".replace(".", ",")


def _old_row_to_data(row) -> Dict[str, Any]:
    return {
        "entrant_id": str(row[0]),
        "full_name": row[1],
        "lang": row[2],
        "subjects": json.loads(row[3]) if row[3] else [],
        "scores": json.loads(row[4]) if row[4] else [],
        "extra_scores": [tuple(x) for x in (json.loads(row[5]) if row[5] else [])],
        "total_ball": row[6],
    }


def _migrate_old_workbook() -> None:
    """Eski JSON-column formatidagi faylni yangi (har biri alohida ustun)
    formatga o'tkazadi. Asl fayl backup sifatida saqlanadi."""
    backup_path = EXCEL_PATH + ".old_json_backup.xlsx"
    shutil.copy2(EXCEL_PATH, backup_path)

    old_wb = load_workbook(EXCEL_PATH, read_only=True)
    old_ws = old_wb["results"] if "results" in old_wb.sheetnames else old_wb.active

    rows_data = []
    for row in old_ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        try:
            data = _old_row_to_data(row)
        except Exception:
            continue
        rows_data.append((data, row[7] if len(row) > 7 else None))
    old_wb.close()

    new_wb = Workbook()
    new_ws = new_wb.active
    new_ws.title = "results"
    new_ws.append(_HEADERS)
    for data, updated_at in rows_data:
        new_row = _row_from_data(data)
        if updated_at:
            new_row[-1] = updated_at
        new_ws.append(new_row)
    new_wb.save(EXCEL_PATH)


def _ensure_workbook() -> None:
    if not os.path.exists(EXCEL_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = "results"
        ws.append(_HEADERS)
        wb.save(EXCEL_PATH)
        return

    wb = load_workbook(EXCEL_PATH, read_only=True)
    ws = wb["results"] if "results" in wb.sheetnames else wb.active
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    wb.close()

    if header == _OLD_HEADERS:
        _migrate_old_workbook()


def _row_from_data(data: Dict[str, Any]) -> list:
    subjects = data.get("subjects") or []
    scores = data.get("scores") or []
    extras = dict(data.get("extra_scores") or [])

    row = [
        str(data.get("entrant_id") or ""),
        data.get("full_name") or "",
        data.get("lang") or "",
    ]

    total_correct = 0
    total_questions = 0
    for i in range(3):
        subj = subjects[i] if i < len(subjects) else {}
        sc = scores[i] if i < len(scores) else {}
        correct = sc.get("correct", 0) if sc else 0
        pct = _fmt_pct((correct / 30) * 100) if sc else ""
        total_correct += correct
        total_questions += 30 if sc else 0
        row += [subj.get("value") or "", correct if sc else "", pct, sc.get("ball") or ""]

    overall_pct = round((total_correct / total_questions) * 100) if total_questions else ""
    row += [total_correct, total_questions, overall_pct]

    for label in _EXTRA_LABELS:
        row.append(extras.get(label, ""))

    row.append(data.get("total_ball") or "")
    row.append(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    return row


def _data_from_row(row) -> Dict[str, Any]:
    def _subject(name, correct, ball):
        if not name and not ball:
            return None, None
        return {"label": "", "value": name or ""}, {"correct": int(correct) if correct not in (None, "") else 0, "ball": ball or ""}

    subjects = []
    scores = []
    for name, correct, _pct, ball in (
        (row[3], row[4], row[5], row[6]),
        (row[7], row[8], row[9], row[10]),
        (row[11], row[12], row[13], row[14]),
    ):
        subj, sc = _subject(name, correct, ball)
        if subj is None:
            continue
        subjects.append(subj)
        scores.append(sc)

    extra_scores = [
        (label, row[18 + i])
        for i, label in enumerate(_EXTRA_LABELS)
        if row[18 + i] not in (None, "")
    ]

    return {
        "entrant_id": str(row[0]),
        "full_name": row[1],
        "lang": row[2],
        "subjects": subjects,
        "scores": scores,
        "extra_scores": extra_scores,
        "total_ball": row[22],
    }


def _sync_lookup(entrant_id: str) -> Optional[Dict[str, Any]]:
    try:
        _ensure_workbook()
        wb = load_workbook(EXCEL_PATH, read_only=True)
        try:
            ws = wb["results"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and str(row[0]) == str(entrant_id):
                    return _data_from_row(row)
        finally:
            wb.close()
    except Exception as e:
        logger.error("[mandat_excel] lookup failed: %r", e)
    return None


def _sync_save(data: Dict[str, Any]) -> None:
    """Excelga saqlashda xatolik bo'lsa ham throw qilmaydi — natijani
    userga berish excel yozuviga bog'liq bo'lmasligi kerak."""
    try:
        _ensure_workbook()
        wb = load_workbook(EXCEL_PATH)
        ws = wb["results"]

        entrant_id = str(data.get("entrant_id") or "")
        target_row_idx = None
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row and str(row[0]) == entrant_id:
                target_row_idx = idx
                break

        new_row = _row_from_data(data)
        if target_row_idx:
            for col, value in enumerate(new_row, start=1):
                ws.cell(row=target_row_idx, column=col, value=value)
        else:
            ws.append(new_row)

        wb.save(EXCEL_PATH)
    except Exception as e:
        logger.error("[mandat_excel] save failed entrant_id=%s: %r", data.get("entrant_id"), e)


async def lookup_cached_result(entrant_id: str) -> Optional[Dict[str, Any]]:
    loop = asyncio.get_event_loop()
    async with _LOCK:
        return await loop.run_in_executor(None, _sync_lookup, entrant_id)


async def save_result_to_cache(data: Dict[str, Any]) -> None:
    loop = asyncio.get_event_loop()
    async with _LOCK:
        await loop.run_in_executor(None, _sync_save, data)


def excel_file_exists() -> bool:
    return os.path.exists(EXCEL_PATH)
